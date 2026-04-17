"""파일 업로드 파서 — Excel(.xlsx/.xls) 및 PDF를 텍스트 블록으로 변환.

용도:
- 대표님이 올린 매물 리스트(Excel) / 계약서·감정평가서(PDF)를
  에이전트가 참조할 수 있는 텍스트로 변환
- 변환 결과는 meeting transcript의 user 메시지로 주입
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".pdf"}
MAX_TEXT_LENGTH = 8000


def parse_file(path: str | Path) -> str:
    """파일을 파싱하여 에이전트 주입용 텍스트 블록으로 반환."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {p}")

    ext = p.suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _parse_excel(p)
    elif ext == ".pdf":
        return _parse_pdf(p)
    else:
        raise ValueError(
            f"지원하지 않는 파일 형식입니다: {ext} "
            f"(지원: {', '.join(sorted(SUPPORTED_EXTENSIONS))})"
        )


def parse_files(paths: list[str | Path]) -> str:
    """여러 파일을 파싱하여 하나의 텍스트 블록으로 합침."""
    blocks: list[str] = []
    for p in paths:
        blocks.append(parse_file(p))
    return "\n\n".join(blocks)


def format_for_agents(file_texts: list[tuple[str, str]]) -> str:
    """(파일명, 파싱결과) 리스트를 에이전트 주입용 블록으로 래핑."""
    if not file_texts:
        return ""
    lines = ["=== 📎 업로드된 파일 데이터 ==="]
    for filename, text in file_texts:
        lines.append(f"\n■ 파일: {filename}")
        lines.append(text)
    lines.append("\n=== 업로드 파일 데이터 끝 ===")
    return "\n".join(lines)


# ------------------------------------------------------------------
# Excel 파싱
# ------------------------------------------------------------------


def _parse_excel(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    blocks: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        if not rows:
            continue

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]

        lines = [f"[시트: {sheet_name}]"]
        lines.append(" | ".join(headers))
        lines.append("-" * min(len(" | ".join(headers)), 80))

        for row in data_rows[:100]:
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))

        if len(data_rows) > 100:
            lines.append(f"... (총 {len(data_rows)}행 중 100행만 표시)")

        lines.append(f"합계: {len(data_rows)}행 × {len(headers)}열")
        blocks.append("\n".join(lines))

    wb.close()

    result = "\n\n".join(blocks)
    if len(result) > MAX_TEXT_LENGTH:
        result = result[:MAX_TEXT_LENGTH] + "\n... (텍스트가 길어 잘림)"
    return result


# ------------------------------------------------------------------
# PDF 파싱
# ------------------------------------------------------------------


def _parse_pdf(path: Path) -> str:
    import pdfplumber

    lines: list[str] = []

    with pdfplumber.open(path) as pdf:
        lines.append(f"[PDF: {path.name}, {len(pdf.pages)}페이지]")

        for i, page in enumerate(pdf.pages[:30], start=1):
            tables = page.extract_tables()
            text = page.extract_text() or ""

            if tables:
                for t_idx, table in enumerate(tables):
                    if not table:
                        continue
                    lines.append(f"\n(p.{i} 표{t_idx + 1})")
                    for row in table:
                        cells = [str(c) if c is not None else "" for c in row]
                        lines.append(" | ".join(cells))
            elif text.strip():
                lines.append(f"\n(p.{i})")
                lines.append(text.strip())

        if len(pdf.pages) > 30:
            lines.append(f"\n... (총 {len(pdf.pages)}페이지 중 30페이지만 표시)")

    result = "\n".join(lines)
    if len(result) > MAX_TEXT_LENGTH:
        result = result[:MAX_TEXT_LENGTH] + "\n... (텍스트가 길어 잘림)"
    return result
